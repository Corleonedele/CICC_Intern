import openpyxl
wb = openpyxl.load_workbook("t.xlsx")

st1 = wb["Sheet1"]
info_st = wb["基本信息表"]


产品基本信息 = ["产品全称","产品简称","产品预备案代码","产品类型",]
产品基本信息dict = {"产品全称":14,"产品简称":15,"产品预备案代码":16,"产品类型":23,}








for i in range(1, st1.max_row+1):

    line_info = str(st1.cell(i, 1).value).strip()

    try:
        index = 产品基本信息.index(line_info)
        index_info = str(st1.cell(i+1, 1).value).strip()

        info_st.cell(产品基本信息dict.get(line_info), 3).value = index_info
        print(line_info, index_info, )   
    except:
        continue

    
wb.save("t1.xlsx")


# with open("t.txt") as f:
#     for i in f:
#         print("\""+i[:-1]+ "\",")