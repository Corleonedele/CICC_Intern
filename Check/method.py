import pandas as pd
from openpyxl.styles import PatternFill
import numpy as np
import os


"""全局参数"""

class Public():

    def initDirs():
        """初始化文件系统
        Fixed_Data 用于存储每日需要用到的固定文件
        Input_Data 用于添加每日需要用到的新增文件
        Output_Data 用于储存程序运行完的文件
        Stored_Data 用于备份各个文件
        """
        if not os.path.exists("./Fixed_Data"):
            os.makedirs("./Fixed_Data") 
        if not os.path.exists("./Input_Data"):
            os.makedirs("./Input_Data") 
        if not os.path.exists("./Output_Data"):
            os.makedirs("./Output_Data") 
        if not os.path.exists("./Fixed_Data/Stored_Data"):
            os.makedirs("./Fixed_Data/Stored_Data") 

    def getFileList(filter_arg=False, path=".", dir_only=False):
        """默认获取当前文件夹下文件, 可设置filter_arg来查找特定文件(eg filter_arg=".py" 会查找python文件, 设置path去查找特定路径下查找文件, 返回该文件夹下所有文件"""
        if filter_arg:
            return list(filter(lambda x: filter_arg in x, os.listdir(path)))
        else:
            if dir_only:
                return list(filter(lambda x: "." not in x, os.listdir(path)))
            return os.listdir(path)
            
    def readFile(file_name) -> str:
        """按文件名查找文件, 如果当前目录没找到, 找且仅找下一级目录, 若仍然没找到则返回None"""
        current_file_list = Public.getFileList()
        if file_name in current_file_list: 
            return os.path.join("./", file_name)
        else:
            for dir in Public.getFileList(dir_only=True):
                if file_name in Public.getFileList(path=dir):
                    return os.path.join("./"+dir, file_name)
        return None



class Compute():

    def maxDrawdown(input_list) -> list:
        """计算最大回撤 返回最大回撤率与最大回撤区间(左, 右)的index"""
        i = np.argmax((np.maximum.accumulate(input_list) - input_list) / np.maximum.accumulate(input_list))
        if i == 0: 
            return 0, 0, 0
        j = np.argmax(input_list[:i])
        return (input_list[j] - input_list[i]) / input_list[j], j, i



class Method():

    def fillWithLastDay(dataFrame, start_col=0, start_row=0) -> pd.DataFrame:
        """把所有0的位置按照上一个数补齐"""
        print("Start Position Col: ",start_col, "Row: ",start_row, "DataFrame Shape:", dataFrame.shape)
        for col in range(start_col, dataFrame.shape[1]):
            tem = 0
            for row in range(start_row, dataFrame.shape[0]):
                try:
                    cell = float(dataFrame.iloc[row, col])
                except:
                    cell = dataFrame.iloc[row, col]
                if cell == 0 or cell == 0.0: 
                    dataFrame.iloc[row, col] = tem
                else: 
                    tem = cell
        return dataFrame


    def mergeSameSheet(file_name, sheet_name, path, **kwargs) -> pd.DataFrame:
        """合并同一个文件夹下多个Excel中的指定sheet"""
        file_list = Public.getFileList(filter_arg=file_name, path=path)
        print("Current Files:", file_list)
        if len(file_list)==0:
            return ValueError
        else:
            return_sheet=pd.DataFrame()

        for file in file_list:
            file_path = path + "/" + file
            read_sheet =  pd.read_excel(file_path, header=kwargs["header"], index_col=kwargs["index_col"], sheet_name=sheet_name)
            return_sheet=pd.concat([return_sheet, read_sheet], ignore_index = False)
        return return_sheet.groupby(return_sheet.index).first() 


    def compareCell(sheet1, sheet2, color_mark, *kwargs):
        """**kwargs格式 start_row, end_row, start_col, end_col
                        row_offset, col_offset
            一共6个参数 """

        start_row, end_row, start_col, end_col, row_offset, col_offset = kwargs[0]
        print("开始比对", "sheet1 start from", (start_row, start_col) ,
                        "sheet2 start from", (start_row+row_offset, start_col+col_offset) ,"\n",
                        "sheet1 end in", (end_row, end_col) ,
                        "sheet2 end in", (start_row+row_offset, start_col+col_offset) ,
            )
        for row in range(start_row, end_row+1):
            for col in range(start_col, end_col+1):
                s1_value = sheet1.cell(row, col).value
                s2_value = sheet2.cell(row+row_offset, col+col_offset)
                if s1_value != s2_value: 
                    if type(s1_value) == float and type(s2_value) == float:
                        if round(s1_value, 5) == round(s2_value, 5):  
                            continue
                    if type(s1_value) == str and type(s2_value) == str:
                        if s1_value.strip() == s2_value.strip(): 
                            continue
                    if color_mark:
                        sheet1.cell(row, col).fill = PatternFill("solid", fgColor="1874CD")
                        sheet2.cell(row+row_offset, col+col_offset).fill = PatternFill("solid", fgColor="FFA300")
                    print("position:","sheet1:", (row, col), "value:", s1_value)
                    print("position:","sheet2:", (row+row_offset, col+col_offset), "value:", s2_value)
        print("比对完成")
        return True
