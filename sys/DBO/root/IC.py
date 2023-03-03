
import openpyxl
from public import Public


class IC():

    path = "./DBO/DB/IC/IC数据测试.xlsx"

    def 比例update():
        MATCH = False
        wb = openpyxl.open(IC.path)
        比例数据st = wb["IC比例数据"]
        交易记录st = wb["交易记录"]
        
        data = Public.readRow(交易记录st, 交易记录st.max_row, 1)
        IC_info = [data[2], data[3], data[37],data[10:30]]
        print(IC_info)
        #  产品代码 	 产品名称 	 IC名称 	 IC代码 	 持有比例 	 产品总份额 
        write_info = [data[2], data[3], 0, 0, 0, 0]



        for row in range(2, 比例数据st.max_row+1):
            row_info = Public.readRow(比例数据st, row, 1)
            # 匹配是否有记录，有则修改，没有则新增
            if IC_info[0] in row_info:
                IC_name = row_info[2]
                for IC_index in [4, 8, 12, 16]:
                    if IC_info[3][IC_index] == None: continue
                    if IC_info[3][IC_index] == IC_name:
                        write_info[2] = IC_info[3][IC_index] # IC名称
                        write_info[4] = float(row_info[4]) * float(row_info[5]) + float(IC_info[3][IC_index+3]) * float(IC_info[2]) / float(row_info[5]) + float(IC_info[2])
                        write_info[5] = float(row_info[4]) + float(IC_info[2])
                        MATCH = True
                        break

        if MATCH:
            Public.writeRow(比例数据st, row, write_info, 1)
            print(write_info)
        else:
            for IC_index in [4, 8, 12, 16]:
                if IC_info[3][IC_index] == None: continue
                write_info[2] = IC_info[3][IC_index] # IC名称
                write_info[5] = float(IC_info[3][IC_index+3]) # IC比例
                write_info[6] = float(IC_info[2]) # 总份额
                Public.writeRow(比例数据st, 比例数据st.max_row+1, write_info, 1) 


        

IC.比例update()