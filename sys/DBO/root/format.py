import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection

REPORT_PATH = "./DBO/DB/REPORT/test_report.xlsx"

YELLOW = PatternFill('solid', fgColor="FFFF00")
GRAY = PatternFill('solid', fgColor="D9D9D9")
BORDER = Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))

def 风险控制指标情况(wb):
    st = wb["风险控制指标情况"]

    # 调整列宽
    st.column_dimensions['A'].width = 5
    st.column_dimensions['B'].width = 40
    st.column_dimensions['C'].width = 20
    st.column_dimensions['D'].width = 15
    st.column_dimensions['E'].width = 25
    st.column_dimensions['F'].width = 20
    st.column_dimensions['G'].width = 20


    # 调整字体 居中
    for row in range(1, st.max_row):
        # 调整行高
        st.row_dimensions[row].height = 30
        for col in range(1, st.max_column+1):
            st.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            st.cell(row, col).border = BORDER
    st.row_dimensions[173].height = 80

    # 调整标题
    st.merge_cells('A1:F1') 
    st.merge_cells('A5:A8') 
    st.merge_cells('A9:A12')
    st.merge_cells('A13:A16')
    st.merge_cells('A17:A20')   
    st.merge_cells('A21:A51')
    st.merge_cells('A52:A172')  

    # 调整颜色
    for col in range(2, 7):
        st.cell(2, col).fill = YELLOW
    for row in [3, 4, 5, 13, 17, 21, 52, 72, 73, 90, 102, 113, 134, 135, 146, 162, 173]: 
        st.cell(row, 2).fill = YELLOW
        

    print("风险控制指标情况 调整结束")

def 备注信息表(wb):
    st = wb["备注信息表"]
    for row in range(1, st.max_row+1):
        # 调整行高
        st.row_dimensions[row].height = 30
        for col in range(1, st.max_column):
            st.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            st.cell(row, col).border = BORDER
    st.column_dimensions['A'].width = 15  
    st.column_dimensions['G'].width = 80    
    print("备注信息表 调整结束")


def 交易记录(wb):
    st = wb["交易记录"]
    for row in range(1, st.max_row+1):
        # 调整行高
        st.row_dimensions[row].height = 30
        for col in range(1, st.max_column):
            st.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            st.cell(row, col).border = BORDER
    st.column_dimensions['A'].width = 15  
    st.column_dimensions['G'].width = 80    


    print("交易记录 调整结束")




def main():

    wb = openpyxl.load_workbook(REPORT_PATH)

    交易记录(wb)
    wb.save(REPORT_PATH)

if __name__ == "__main__":
    main()