import openpyxl


STANDARD_DATE = "20230203"
class Path():
    交易数据_OLD = "./DBO/DB/UPDATED/交易记录.xlsx"
    交易数据 = "./DBO/DB/TEST/交易记录.xlsx"
    私募种子基金持仓日报表_OLD = "./DBO/DB/TEST/私募种子基金持仓日报表.xlsx"
    私募种子基金持仓日报表 = "./DBO/DB/TEST/私募种子基金持仓日报表.xlsx"



def update(row_data): # 修改私募种子基金持仓日报表中的底层资产私募配置情况
    row_index = 0
    print(row_data)
    私募种子基金持仓日报表 = openpyxl.load_workbook(Path.私募种子基金持仓日报表)["Sheet"]
    交易记录 = openpyxl.load_workbook(Path.交易数据)["Sheet"]
    底层资产私募配置情况 = (14, 2)
    私募种子基金业务资产及盈亏情况 = (3, 2)
    策略分布 = (4, 10)
    FOF单一资管计划 = (4, 14)

    for row in range(14, 私募种子基金持仓日报表.max_row+1):
        if row_data[2] == 私募种子基金持仓日报表.cell(row, 3).value:
            row_index = row
            break

    





def run():
    modified = ["交易数据"]
    # 交易数据更新
    # 生成持仓日报表
    # 净值更新
    # 生成风险日报表

    if "交易数据" in modified:
        交易数据 = openpyxl.load_workbook(Path.交易数据)["Sheet"]
        交易数据_OLD = openpyxl.load_workbook(Path.交易数据_OLD)["Sheet"]

        if 交易数据_OLD.max_row - 交易数据.max_row == 0: 
            print("交易记录表更新未检测到")
            return
        else:
            print("交易记录表更新", 交易数据.max_row - 交易数据_OLD.max_row , "条")
            start_index = 交易数据_OLD.max_row
            end_index = 交易数据.max_row
                    

                # print(交易数据.cell(trade_record, 2).value)
            # 修改底层私募
            # for line in 私募种子基金持仓日报表



    # 清空TEST文件夹
    # 复制UPDATED文件到TEST文件夹
