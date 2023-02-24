import openpyxl
from public import Public


class 风险控制指标情况():

    path = "./DBO/DB/TEST/风险控制指标情况.xlsx"


class 私募种子基金持仓日报表():
    path = "./DBO/DB/TEST/私募种子基金持仓日报表.xlsx"


class 交易记录():
    path = "./DBO/DB/TEST/交易记录.xlsx"

    def 公用输入():
        return 0

    def 加权成本(sheet):
        产品清算金额汇总 = 0
        产品持仓数量汇总 = 0

        for row in range(1+1, sheet.max_row+1):
            direction = sheet.cell(row, 2).value
            value = float(sheet.cell(row, 20).value)
            amount = float(sheet.cell(row, 15).value)

            if direction == "分红再投":
                产品清算金额汇总 += value
            elif direction == "现金分红":
                产品清算金额汇总 -= value
            elif direction == "调减":
                产品清算金额汇总 += value
            elif direction == "追加":
                产品清算金额汇总 -= value
            elif direction == "赎回":
                产品清算金额汇总 -= value


            if direction == "分红再投":
                产品持仓数量汇总 += amount
            elif direction == "现金分红":
                产品持仓数量汇总 += amount
            elif direction == "调减":
                产品持仓数量汇总 -= amount
            elif direction == "追加":
                产品持仓数量汇总 += amount
            elif direction == "赎回":
                产品持仓数量汇总 -= amount

        if 产品持仓数量汇总 == 0:
            return 0
        return 产品清算金额汇总 / 产品持仓数量汇总

    def 本年度加权成本(sheet):
        产品本年度成本汇总 = 0
        产品持仓数量变动汇总  = 0

        for row in range(1+1, sheet.max_row):
            direction = sheet.cell(row, 2).value
            value = float(sheet.cell(row, 25).value)
            amount = float(sheet.cell(row, 15).value)

            if direction == "分红再投":
                产品本年度成本汇总 += value
            elif direction == "现金分红":
                产品本年度成本汇总 -= value
            elif direction == "调减":
                产品本年度成本汇总 += value
            elif direction == "追加":
                产品本年度成本汇总 -= value
            elif direction == "赎回":
                产品本年度成本汇总 -= value


            if direction == "分红再投":
                产品持仓数量变动汇总 += amount
            elif direction == "现金分红":
                产品持仓数量变动汇总 += amount
            elif direction == "调减":
                产品持仓数量变动汇总 -= amount
            elif direction == "追加":
                产品持仓数量变动汇总 += amount
            elif direction == "赎回":
                产品持仓数量变动汇总 -= amount

        if 产品持仓数量变动汇总 == 0:
            return 0
        return 产品本年度成本汇总 / 产品持仓数量变动汇总


    def 追加(body_dict):
        wb = openpyxl.load_workbook(交易记录.path)
        st = wb["Sheet"]
        write_row = st.max_row + 1


        # 手动输入
        # 成交时间 col = 1
        st.cell(write_row, 1).value = body_dict["成交时间"]
        # 买卖方向 col = 2
        st.cell(write_row, 2).value = "追加"
        # 证劵代码 col = 3
        st.cell(write_row, 3).value = body_dict["证劵代码"]
        # 产品名称 col = 4
        st.cell(write_row, 4).value = body_dict["产品名称"]
        # 产品管理人 col = 5
        st.cell(write_row, 5).value = body_dict["产品管理人"]
        # 策略类型 col = 6
        st.cell(write_row, 6).value = body_dict["策略类型"]
        # 策略类型_新 col = 7
        st.cell(write_row, 7).value = body_dict["策略类型_新"]
        # 跟踪指数 col = 8
        st.cell(write_row, 8).value = body_dict["跟踪指数"]
        # 细分策略 col = 9
        st.cell(write_row, 9).value = body_dict["细分策略"]
        # 产品分类 col = 10
        st.cell(write_row, 10).value = body_dict["产品分类"]
        # 初始投资金额 col = 14
        st.cell(write_row, 14).value = body_dict["初始投资金额"]
        # 成交数量 col = 15
        st.cell(write_row, 15).value = body_dict["成交数量"]
        # 成交金额_万元 col = 16
        st.cell(write_row, 16).value = body_dict["成交金额_万元"]
        # 本年度成本价 col = 24
        st.cell(write_row, 24).value = body_dict["本年度成本价"]
        # 分支机构 col =11
        st.cell(write_row, 11).value = body_dict["分支机构"]
        # 推荐IC col = 12
        st.cell(write_row, 12).value = body_dict["推荐IC"]
        # 考核承担IC col = 13
        st.cell(write_row, 13).value = body_dict["考核承担IC"]
        # IC分摊比例 = 17
        st.cell(write_row, 17).value = body_dict["IC分摊比例"]


        # 公式计算

        # 初始成交价 col = 18
        st.cell(write_row, 18).value = float(body_dict["成交金额_万元"]) / float(body_dict["成交数量"])
        # 数量变动 col = 19
        st.cell(write_row, 19).value = float(body_dict["成交数量"])
        # 清算金额 col = 20
        st.cell(write_row, 20).value = (-1) * float(body_dict["成交金额_万元"])
        # 持仓金额变动 col = 21
        st.cell(write_row, 21).value = (-1) * float(body_dict["成交金额_万元"])
        # 加权成本 col = 22
        st.cell(write_row, 22).value = 交易记录.加权成本(st)
        # 卖出成本 col = 23
        st.cell(write_row, 23).value = 0
        # 本年度成本 col = 25
        st.cell(write_row, 25).value = 0
        # 本年度加权成本 col = 26
        st.cell(write_row, 26).value = 交易记录.本年度加权成本(st)
        # 本年度卖出成本 col = 27
        st.cell(write_row, 27).value = 0
        # 累计已实现收益 col = 28
        st.cell(write_row, 28).value = 0
        # 本年度已实现收益 col = 29
        st.cell(write_row, 29).value = 0

        wb.save(交易记录.path)


    def 调减(body_dict):
        wb = openpyxl.load_workbook(交易记录.path)
        st = wb["Sheet1"]
        write_row = st.max_row + 1

        # 手动输入
        # 成交时间 col = 1
        st.cell(write_row, 1).value = body_dict["成交时间"]
        # 买卖方向 col = 2
        st.cell(write_row, 2).value = "调减"
        # 证劵代码 col = 3
        st.cell(write_row, 3).value = body_dict["证劵代码"]
        # 产品名称 col = 4
        st.cell(write_row, 4).value = body_dict["产品名称"]
        # 产品管理人 col = 5
        st.cell(write_row, 5).value = body_dict["产品管理人"]
        # 策略类型 col = 6
        st.cell(write_row, 6).value = body_dict["策略类型"]
        # 策略类型_新 col = 7
        st.cell(write_row, 7).value = body_dict["策略类型_新"]
        # 跟踪指数 col = 8
        st.cell(write_row, 8).value = body_dict["跟踪指数"]
        # 细分策略 col = 9
        st.cell(write_row, 9).value = body_dict["细分策略"]
        # 产品分类 col = 10
        st.cell(write_row, 10).value = body_dict["产品分类"]
        # 初始投资金额 col = 14
        st.cell(write_row, 14).value = body_dict["初始投资金额"]
        # 成交数量 col = 15
        st.cell(write_row, 15).value = body_dict["成交数量"]
        # 成交金额_万元 col = 16
        st.cell(write_row, 16).value = body_dict["成交金额_万元"]
        # 本年度成本价 col = 24
        st.cell(write_row, 24).value = body_dict["本年度成本价"]
        # 分支机构 col =11
        st.cell(write_row, 11).value = body_dict["分支机构"]
        # 推荐IC col = 12
        st.cell(write_row, 12).value = body_dict["推荐IC"]
        # 考核承担IC col = 13
        st.cell(write_row, 13).value = body_dict["考核承担IC"]
        # IC分摊比例 = 17
        st.cell(write_row, 17).value = body_dict["IC分摊比例"]


        # 公式计算

        # 初始成交价 col = 18
        st.cell(write_row, 18).value = float(body_dict["成交金额_万元"]) / float(body_dict["成交数量"])
        # 数量变动 col = 19
        st.cell(write_row, 19).value = float(body_dict["成交数量"])
        # 清算金额 col = 20
        st.cell(write_row, 20).value = (-1) * float(body_dict["成交金额_万元"])
        # 持仓金额变动 col = 21
        st.cell(write_row, 21).value = (-1) * float(body_dict["成交金额_万元"])
        # 加权成本 col = 22
        st.cell(write_row, 22).value = 交易记录.加权成本(st)
        # 卖出成本 col = 23
        st.cell(write_row, 23).value = 0
        # 本年度成本 col = 25
        st.cell(write_row, 25).value = 0
        # 本年度加权成本 col = 26
        st.cell(write_row, 26).value = 交易记录.本年度加权成本(st)
        # 本年度卖出成本 col = 27
        st.cell(write_row, 27).value = 0
        # 累计已实现收益 col = 28
        st.cell(write_row, 28).value = 0
        # 本年度已实现收益 col = 29
        st.cell(write_row, 29).value = 0


        wb.save(交易记录.path)

    def 赎回(body_dict):
        wb = openpyxl.load_workbook(交易记录.path)
        st = wb["Sheet1"]
        write_row = st.max_row + 1

        # 手动输入
        # 成交时间 col = 1
        st.cell(write_row, 1).value = body_dict["成交时间"]
        # 买卖方向 col = 2
        st.cell(write_row, 2).value = "赎回"
        # 证劵代码 col = 3
        st.cell(write_row, 3).value = body_dict["证劵代码"]
        # 产品名称 col = 4
        st.cell(write_row, 4).value = body_dict["产品名称"]
        # 产品管理人 col = 5
        st.cell(write_row, 5).value = body_dict["产品管理人"]
        # 策略类型 col = 6
        st.cell(write_row, 6).value = body_dict["策略类型"]
        # 策略类型_新 col = 7
        st.cell(write_row, 7).value = body_dict["策略类型_新"]
        # 跟踪指数 col = 8
        st.cell(write_row, 8).value = body_dict["跟踪指数"]
        # 细分策略 col = 9
        st.cell(write_row, 9).value = body_dict["细分策略"]
        # 产品分类 col = 10
        st.cell(write_row, 10).value = body_dict["产品分类"]
        # 初始投资金额 col = 14
        st.cell(write_row, 14).value = body_dict["初始投资金额"]
        # 成交数量 col = 15
        st.cell(write_row, 15).value = body_dict["成交数量"]
        # 成交金额_万元 col = 16
        st.cell(write_row, 16).value = body_dict["成交金额_万元"]
        # 本年度成本价 col = 24
        st.cell(write_row, 24).value = body_dict["本年度成本价"]
        # 分支机构 col =11
        st.cell(write_row, 11).value = body_dict["分支机构"]
        # 推荐IC col = 12
        st.cell(write_row, 12).value = body_dict["推荐IC"]
        # 考核承担IC col = 13
        st.cell(write_row, 13).value = body_dict["考核承担IC"]
        # IC分摊比例 = 17
        st.cell(write_row, 17).value = body_dict["IC分摊比例"]


        # 公式计算

        # 初始成交价 col = 18
        st.cell(write_row, 18).value = float(body_dict["成交金额_万元"]) / float(body_dict["成交数量"])
        # 数量变动 col = 19
        st.cell(write_row, 19).value = float(body_dict["成交数量"])
        # 清算金额 col = 20
        st.cell(write_row, 20).value = (-1) * float(body_dict["成交金额_万元"])
        # 持仓金额变动 col = 21
        st.cell(write_row, 21).value = (-1) * float(body_dict["成交金额_万元"])
        # 加权成本 col = 22
        st.cell(write_row, 22).value = 交易记录.加权成本(st)
        # 卖出成本 col = 23
        st.cell(write_row, 23).value = 0
        # 本年度成本 col = 25
        st.cell(write_row, 25).value = 0
        # 本年度加权成本 col = 26
        st.cell(write_row, 26).value = 交易记录.本年度加权成本(st)
        # 本年度卖出成本 col = 27
        st.cell(write_row, 27).value = 0
        # 累计已实现收益 col = 28
        st.cell(write_row, 28).value = 0
        # 本年度已实现收益 col = 29
        st.cell(write_row, 29).value = 0


        wb.save(交易记录.path)

    def 现金分红(body_dict):
        wb = openpyxl.load_workbook(交易记录.path)
        st = wb["Sheet"]
        write_row = st.max_row + 1

        # 手动输入
        # 成交时间 col = 1
        st.cell(write_row, 1).value = body_dict["成交时间"]
        # 买卖方向 col = 2
        st.cell(write_row, 2).value = "现金分红"
        # 证劵代码 col = 3
        st.cell(write_row, 3).value = body_dict["证劵代码"]
        # 产品名称 col = 4
        st.cell(write_row, 4).value = body_dict["产品名称"]
        # 产品管理人 col = 5
        st.cell(write_row, 5).value = body_dict["产品管理人"]
        # 策略类型 col = 6
        st.cell(write_row, 6).value = body_dict["策略类型"]
        # 策略类型_新 col = 7
        st.cell(write_row, 7).value = body_dict["策略类型_新"]
        # 跟踪指数 col = 8
        st.cell(write_row, 8).value = body_dict["跟踪指数"]
        # 细分策略 col = 9
        st.cell(write_row, 9).value = body_dict["细分策略"]
        # 产品分类 col = 10
        st.cell(write_row, 10).value = body_dict["产品分类"]
        # 初始投资金额 col = 14
        st.cell(write_row, 14).value = body_dict["初始投资金额"]
        # 成交数量 col = 15
        st.cell(write_row, 15).value = body_dict["成交数量"]
        # 成交金额_万元 col = 16
        st.cell(write_row, 16).value = body_dict["成交金额_万元"]
        # 本年度成本价 col = 24
        st.cell(write_row, 24).value = body_dict["本年度成本价"]
        # 分支机构 col =11
        st.cell(write_row, 11).value = body_dict["分支机构"]
        # 推荐IC col = 12
        st.cell(write_row, 12).value = body_dict["推荐IC"]
        # 考核承担IC col = 13
        st.cell(write_row, 13).value = body_dict["考核承担IC"]
        # IC分摊比例 = 17
        st.cell(write_row, 17).value = body_dict["IC分摊比例"]


        # 公式计算

        # 初始成交价 col = 18
        st.cell(write_row, 18).value = float(body_dict["成交金额_万元"]) / float(body_dict["成交数量"])
        # 数量变动 col = 19
        st.cell(write_row, 19).value = float(body_dict["成交数量"])
        # 清算金额 col = 20
        st.cell(write_row, 20).value = (-1) * float(body_dict["成交金额_万元"])
        # 持仓金额变动 col = 21
        st.cell(write_row, 21).value = (-1) * float(body_dict["成交金额_万元"])
        # 加权成本 col = 22
        st.cell(write_row, 22).value = 交易记录.加权成本(st)
        # 卖出成本 col = 23
        st.cell(write_row, 23).value = 0
        # 本年度成本 col = 25
        st.cell(write_row, 25).value = 0
        # 本年度加权成本 col = 26
        st.cell(write_row, 26).value = 交易记录.本年度加权成本(st)
        # 本年度卖出成本 col = 27
        st.cell(write_row, 27).value = 0
        # 累计已实现收益 col = 28
        st.cell(write_row, 28).value = 0
        # 本年度已实现收益 col = 29
        st.cell(write_row, 29).value = 0


        wb.save(交易记录.path)

    def 分红再投(body_dict):
        wb = openpyxl.load_workbook(交易记录.path)
        st = wb["Sheet"]
        write_row = st.max_row + 1

        # 手动输入
        # 成交时间 col = 1
        st.cell(write_row, 1).value = body_dict["成交时间"]
        # 买卖方向 col = 2
        st.cell(write_row, 2).value = "分红再投"
        # 证劵代码 col = 3
        st.cell(write_row, 3).value = body_dict["证劵代码"]
        # 产品名称 col = 4
        st.cell(write_row, 4).value = body_dict["产品名称"]
        # 产品管理人 col = 5
        st.cell(write_row, 5).value = body_dict["产品管理人"]
        # 策略类型 col = 6
        st.cell(write_row, 6).value = body_dict["策略类型"]
        # 策略类型_新 col = 7
        st.cell(write_row, 7).value = body_dict["策略类型_新"]
        # 跟踪指数 col = 8
        st.cell(write_row, 8).value = body_dict["跟踪指数"]
        # 细分策略 col = 9
        st.cell(write_row, 9).value = body_dict["细分策略"]
        # 产品分类 col = 10
        st.cell(write_row, 10).value = body_dict["产品分类"]
        # 初始投资金额 col = 14
        st.cell(write_row, 14).value = body_dict["初始投资金额"]
        # 成交数量 col = 15
        st.cell(write_row, 15).value = body_dict["成交数量"]
        # 成交金额_万元 col = 16
        st.cell(write_row, 16).value = body_dict["成交金额_万元"]
        # 本年度成本价 col = 24
        st.cell(write_row, 24).value = body_dict["本年度成本价"]
        # 分支机构 col =11
        st.cell(write_row, 11).value = body_dict["分支机构"]
        # 推荐IC col = 12
        st.cell(write_row, 12).value = body_dict["推荐IC"]
        # 考核承担IC col = 13
        st.cell(write_row, 13).value = body_dict["考核承担IC"]
        # IC分摊比例 = 17
        st.cell(write_row, 17).value = body_dict["IC分摊比例"]


        # 公式计算

        # 初始成交价 col = 18
        st.cell(write_row, 18).value = float(body_dict["成交金额_万元"]) / float(body_dict["成交数量"])
        # 数量变动 col = 19
        st.cell(write_row, 19).value = float(body_dict["成交数量"])
        # 清算金额 col = 20
        st.cell(write_row, 20).value = (-1) * float(body_dict["成交金额_万元"])
        # 持仓金额变动 col = 21
        st.cell(write_row, 21).value = (-1) * float(body_dict["成交金额_万元"])
        # 加权成本 col = 22
        st.cell(write_row, 22).value = 交易记录.加权成本(st)
        # 卖出成本 col = 23
        st.cell(write_row, 23).value = 0
        # 本年度成本 col = 25
        st.cell(write_row, 25).value = 0
        # 本年度加权成本 col = 26
        st.cell(write_row, 26).value = 交易记录.本年度加权成本(st)
        # 本年度卖出成本 col = 27
        st.cell(write_row, 27).value = 0
        # 累计已实现收益 col = 28
        st.cell(write_row, 28).value = 0
        # 本年度已实现收益 col = 29
        st.cell(write_row, 29).value = 0


        wb.save(交易记录.path)


class 私募种子基金业务资产及盈亏情况(私募种子基金持仓日报表):
    def 可用资金额度():
        pos = (5, 3)
    def 累计投资金额():
        pos = (6, 3)
    def 一对一盈亏():
        pos = (7, 3)
    def 一对一持仓市值():
        pos = (8, 3)
    def 持仓总市值():
        pos = (9, 3)
    def 可用资金额度_批准资金额度():
        pos = (5, 5)
    def 总盈亏比例():
        pos = (6, 5)
    def 一对多盈亏():
        pos = (7, 5)
    def 总盈亏():
        pos = (8, 5)
    def 一对多持仓市值():
        pos = (9, 5)
    def 已实现盈亏():
        pos = (10 ,5)

class 策略分布(私募种子基金持仓日报表):
    def 股票多头策略():
        pos = (5, 11) # 持仓市值 市值占比col+1
    def 股票量化策略_指数增强():
        pos = (6, 11)
    def 股票量化策略_空气指增():
        pos = (7, 11)
    def 股票量化策略_量化择时():
        pos = (8, 11)
    def 量化对冲():
        pos = (9, 11)
    def 多策略灵活():
        pos = (10, 11)
    def 宏观对冲():
        pos = (11, 11)
    def 量化期货():
        pos = (12, 11)

class 金富一号(私募种子基金持仓日报表):
    def FOF单一资管计划():
        pos = (5, 14)
    def 委托财产():
        pos = (6, 14)
    def 持有份额():
        pos = (7, 14)
    def 追加资产():
        pos = (8, 14)
    def 单位净值():
        pos = (9, 14)
    def 当前资产净值():
        pos = (10, 14)
    def 盈亏():
        pos = (11, 14)

class 底层资产私募配置情况(私募种子基金持仓日报表):

    def 产品代码(交易记录st):
        # 根据产品分类读取产品代码
        code_type = Public.readColumn(交易记录st, 10, 1)
        code = Public.readColumn(交易记录st, 3, 1)
        一对一产品_list_tem = []
        一对多产品_list_tem = []
        for index, i in enumerate(code_type):
            if i.strip() == "一对一产品":
                一对一产品_list_tem.append((code[index], index+1)) # 加一是因为表头
            elif i.strip() == "一对多产品":
                一对多产品_list_tem.append((code[index], index+1))
        一对一产品_list_tem.reverse()
        一对多产品_list_tem.reverse()
        tem_1 = []
        tem_2 = []
        for i in 一对一产品_list_tem:
            if i[0] not in tem_2:
                tem_1.append(i)
                tem_2.append(i[0])
        tem_3 = []
        tem_4 = []
        for i in 一对多产品_list_tem:
            if i[0] not in tem_4:
                tem_3.append(i)
                tem_4.append(i[0])     

        return tem_1, tem_3

    def run():
        start_row = 15
        start_col = 2
        交易记录st = openpyxl.load_workbook(交易记录.path)["Sheet"]
        底层资产私募配置情况st = openpyxl.load_workbook(私募种子基金持仓日报表.path)["Sheet"]
    
        一对一产品, 一对多产品 = 底层资产私募配置情况.产品代码(交易记录st)
        
        test_wb = openpyxl.Workbook()
        test_st = test_wb.active
        一对一产品code = [i[0] for i in 一对一产品]
        一对一产品index = [i[1] for i in 一对一产品]
        一对多产品code = [i[0] for i in 一对多产品]
        一对多产品index = [i[1] for i in 一对多产品]
        Public.writeColumn(test_st, 2, 一对一产品code, 1)
        Public.writeColumn(test_st, 2, 一对多产品code, len(一对一产品)+1)

        # 产品类型
        私募基金产品类型 = []
        for row_index in 一对一产品index + 一对多产品index:
            私募基金产品类型.append(交易记录st.cell(row_index, 10).value)
        Public.writeColumn(test_st, 1, 私募基金产品类型, 1)
        

        # 产品名称
        产品名称 = []
        for row_index in 一对一产品index + 一对多产品index:
            产品名称.append(交易记录st.cell(row_index, 4).value)
        Public.writeColumn(test_st, 3, 产品名称, 1)

        # 产品管理人
        产品管理人 = []
        for row_index in 一对一产品index + 一对多产品index:
            产品管理人.append(交易记录st.cell(row_index, 5).value)
        Public.writeColumn(test_st, 4, 产品管理人, 1)

        # 策略类型
        策略类型 = []
        for row_index in 一对一产品index + 一对多产品index:
            策略类型.append(交易记录st.cell(row_index, 6).value)
        Public.writeColumn(test_st, 5, 策略类型, 1)

        # 策略类型_新
        策略类型_新 = []
        for row_index in 一对一产品index + 一对多产品index:
            策略类型_新.append(交易记录st.cell(row_index, 7).value)
        Public.writeColumn(test_st, 6, 策略类型, 1)

        # 对标指数
        对标指数 = []
        for row_index in 一对一产品index + 一对多产品index:
            对标指数.append(交易记录st.cell(row_index, 8).value)
        Public.writeColumn(test_st, 7, 对标指数, 1)

        # 细分策略
        细分策略 = []
        for row_index in 一对一产品index + 一对多产品index:
            细分策略.append(交易记录st.cell(row_index, 9).value)
        Public.writeColumn(test_st, 8, 细分策略, 1)
        
        # 投资金额（万元）sumif条件求和
        
        # 持有份额 sumif条件求和

        # 初始成本价
        初始成本价_元 = []
        for row_index in 一对一产品index + 一对多产品index:
            初始成本价_元.append(交易记录st.cell(row_index, 23).value)
        Public.writeColumn(test_st, 12, 初始成本价_元, 1)

        # 本年度成本价
        本年度成本价_元 = []
        for row_index in 一对一产品index + 一对多产品index:
            本年度成本价_元.append(交易记录st.cell(row_index, 27).value)
        Public.writeColumn(test_st, 13, 本年度成本价_元, 1)

        # 单位净值 来自RPA取数

        #





        




        test_wb.save("test.xlsx")
        print("Update Done")


底层资产私募配置情况.run()




