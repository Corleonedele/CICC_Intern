import openpyxl
import time

DB_PATH = "./DBO/DB/"
TEMPLATE_PATH = "./DBO/DB/TEMPLATE/"
REPORT_PATH = "./DBO/DB/REPORT/"
TEST_PATH = "./DBO/DB/TEST/"





def sheet_copy_paste(to_st, from_st):
    for row in range(1, from_st.max_row+1):
        for col in range(1, from_st.max_column+1):
            to_st.cell(row, col).value = from_st.cell(row, col).value


def init_empty_template():
    sheet_name = ["风险控制指标情况", "备注信息表", "私募种子基金持仓日报表", "交易记录", "份额-产品到期日期", "股票多头", "指数增强", "空气指增", "量化择时", "量化对冲", "宏观对冲", "量化期货", "多策略灵活配置", "资金流水", "年初资产+追加资产", "Mapping", "FOF(YTD)"]
    for i in sheet_name:
        wb = openpyxl.Workbook()
        wb.active
        wb.save(TEMPLATE_PATH+i+".xlsx")


def merge_report():
    sheet_name = ["备注信息表", "私募种子基金持仓日报表", "交易记录", "份额-产品到期日期", "股票多头", "指数增强", "空气指增", "量化择时", "量化对冲", "宏观对冲", "量化期货", "多策略灵活配置", "资金流水", "年初资产+追加资产", "Mapping", "FOF(YTD)"]
    report = openpyxl.Workbook()
    st = report.active
    st.title = "风险控制指标情况"
    for i in sheet_name:
        report.create_sheet(i)

    for name in sheet_name+["风险控制指标情况"]:
        wb = openpyxl.load_workbook(TEST_PATH+name+".xlsx")
        st_from = wb["Sheet"]
        st_to = report[name]
        sheet_copy_paste(st_to, st_from)
        print(name, "loading")

    report.save(REPORT_PATH+"test_report.xlsx")
    print("日报已生成")




start_time = time.time()

merge_report()

end_time = time.time()

print("Time:", end_time-start_time)